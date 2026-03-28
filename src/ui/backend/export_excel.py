"""
export_excel.py — Excel export for the Automated Budgeting API.

GET /api/transactions/export?month=YYYY-MM

Three-sheet workbook:
  1. Monthly Budget   — template-style layout; categories come from categories.json
                        so it automatically reflects any additions/removals.
                        Budget column is blank for the user to fill in;
                        Actual is populated from the DB;
                        Remaining = Budget − Actual (live formula).
  2. Transactions     — date-sorted ledger: Date, Type, Description, Category,
                        Amount, Notes  (no Frequency column).
  3. Category Detail  — total / count / avg per category, % of expenses, % of income.
"""

import io
import json as _json
import logging
from pathlib import Path

from fastapi import APIRouter
from fastapi.responses import JSONResponse, StreamingResponse

from src.ui.backend.deps import _DB_AVAILABLE, get_engine

router = APIRouter()

# categories.json lives at <project_root>/config/categories.json
# This file: src/ui/backend/export_excel.py → parents[3] = project root
_CONF_ROOT = Path(__file__).parents[3] / 'config'


def _load_categories() -> tuple[list[str], dict[str, list[str]]]:
    """Return (categories_list, subcategory_map) from categories.json."""
    try:
        with open(_CONF_ROOT / 'categories.json') as fh:
            data = _json.load(fh)
        return data.get('categories', []), data.get('subcategories', {})
    except Exception:
        return [], {}


@router.get("/api/transactions/export")
def export_transactions(month: str = ''):
    """Export a formatted Excel workbook for the given month (or all months)."""
    if not _DB_AVAILABLE:
        return JSONResponse(status_code=503, content={'error': 'Database unavailable'})

    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
        from sqlalchemy import text as _text

        # ── Load categories (dynamic, from categories.json) ──────────────────
        all_cats, subcats = _load_categories()
        child_set = {c for children in subcats.values() for c in children}
        top_level = [c for c in all_cats if c not in child_set]

        eng = get_engine()
        params: dict = {}
        month_filter = (
            " AND SUBSTR(tx_date,7,4)||'-'||SUBSTR(tx_date,1,2) = :month"
            if month else ""
        )
        if month:
            params['month'] = month

        with eng.connect() as conn:
            tx_rows = conn.execute(_text(
                "SELECT tx_hash, tx_date, place, amount, category, label, tx_type "
                "FROM transactions WHERE tx_type != 'transfer'" + month_filter +
                " ORDER BY tx_date ASC, place ASC"
            ), params).fetchall()

        label_month = month if month else 'All Months'

        # ── Style helpers ─────────────────────────────────────────────────────
        def _fill(hex6: str) -> PatternFill:
            return PatternFill('solid', fgColor=hex6)

        def _fnt(bold=False, color='000000', size=11, italic=False) -> Font:
            return Font(bold=bold, color=color, size=size, italic=italic)

        def _aln(h='left', v='center', wrap=False) -> Alignment:
            return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

        C_NAVY  = '1E3A5F'
        C_BLUE  = '2D6A9F'
        C_LBLUE = 'D0E4F7'
        C_GREY  = 'F8FAFC'
        C_WHITE = 'FFFFFF'
        C_GREEN = 'F0FDF4'
        C_RED   = 'FEF2F2'
        C_PURP  = 'F5F3FF'
        C_TBLUE = 'EBF3FB'
        MONEY   = '#,##0.00'
        PCT     = '0.0%'

        # ── Pre-compute totals ────────────────────────────────────────────────
        total_income = sum(
            float(r[3] or 0) for r in tx_rows
            if r[6] == 'income' and str(r[5] or '').lower() != 'investment_transfer'
        )
        total_expenses = sum(
            float(r[3] or 0) for r in tx_rows
            if r[6] == 'expense' and float(r[3] or 0) >= 0
        )
        total_reimb = sum(
            abs(float(r[3] or 0)) for r in tx_rows
            if r[6] == 'expense' and float(r[3] or 0) < 0
        )
        net = total_income - total_expenses + total_reimb

        cat_actuals: dict[str, float] = {}
        for r in tx_rows:
            if r[6] != 'expense' or float(r[3] or 0) < 0:
                continue
            cat = str(r[4] or 'Uncategorized').strip()
            cat_actuals[cat] = cat_actuals.get(cat, 0.0) + float(r[3] or 0)

        income_salary = sum(
            float(r[3] or 0) for r in tx_rows
            if r[6] == 'income'
            and str(r[5] or '').lower() not in ('bonus', 'reimbursement', 'investment_transfer')
        )
        income_bonus = sum(
            float(r[3] or 0) for r in tx_rows
            if r[6] == 'income' and str(r[5] or '').lower() == 'bonus'
        )

        wb = openpyxl.Workbook()

        # ════════════════════════════════════════════════════════════════════
        # SHEET 1 — Monthly Budget  (A=Item | B=Budget | C=Actual | D=Remaining)
        # ════════════════════════════════════════════════════════════════════
        ws1 = wb.active
        ws1.title = 'Monthly Budget'

        for col, w in zip('ABCD', [34, 16, 16, 16]):
            ws1.column_dimensions[col].width = w

        brow = 1  # running row cursor for this sheet

        # Title banner
        ws1.cell(brow, 1, f'Personal Monthly Budget — {label_month}')
        ws1.cell(brow, 1).fill      = _fill(C_NAVY)
        ws1.cell(brow, 1).font      = _fnt(bold=True, color=C_WHITE, size=14)
        ws1.cell(brow, 1).alignment = _aln('center')
        ws1.merge_cells(f'A{brow}:D{brow}')
        ws1.row_dimensions[brow].height = 30
        brow += 1

        # Hint line
        ws1.cell(brow, 1,
                 '← Enter your targets in the "Budget" column. '
                 'Remaining = Budget − Actual (updates automatically).')
        ws1.cell(brow, 1).font      = _fnt(italic=True, size=9, color='555555')
        ws1.cell(brow, 1).alignment = _aln()
        ws1.merge_cells(f'A{brow}:D{brow}')
        ws1.row_dimensions[brow].height = 14
        brow += 1

        ws1.freeze_panes = f'A{brow + 2}'

        def _sec(label: str):
            nonlocal brow
            ws1.cell(brow, 1, label)
            ws1.cell(brow, 1).fill      = _fill(C_BLUE)
            ws1.cell(brow, 1).font      = _fnt(bold=True, color=C_WHITE, size=11)
            ws1.cell(brow, 1).alignment = _aln()
            ws1.merge_cells(f'A{brow}:D{brow}')
            ws1.row_dimensions[brow].height = 20
            brow += 1

        def _col_hdrs():
            nonlocal brow
            for ci, hdr in enumerate(['Item', 'Budget', 'Actual', 'Remaining'], 1):
                c = ws1.cell(brow, ci, hdr)
                c.fill      = _fill(C_LBLUE)
                c.font      = _fnt(bold=True, size=10)
                c.alignment = _aln('center' if ci > 1 else 'left')
            ws1.row_dimensions[brow].height = 18
            brow += 1

        def _brow_data(label: str, actual: float,
                       indent=False, is_total=False, bg: str | None = None):
            nonlocal brow
            txt = ('    ' if indent else '') + label
            if is_total:
                bg_c = C_GREEN if 'INCOME' in label else C_RED
            else:
                bg_c = bg or (C_GREY if brow % 2 == 0 else C_WHITE)
            ws1.cell(brow, 1, txt)
            ws1.cell(brow, 2, '')
            ws1.cell(brow, 3, round(actual, 2))
            ws1.cell(brow, 4, f'=B{brow}-C{brow}')
            for ci in range(1, 5):
                c = ws1.cell(brow, ci)
                c.fill      = _fill(bg_c)
                c.alignment = _aln('right' if ci > 1 else 'left')
            for ci in (2, 3, 4):
                ws1.cell(brow, ci).number_format = MONEY
            if is_total:
                for ci in range(1, 5):
                    ws1.cell(brow, ci).font = _fnt(bold=True, size=11)
            elif indent:
                ws1.cell(brow, 1).font = _fnt(italic=True, size=10)
            ws1.row_dimensions[brow].height = 17
            brow += 1

        # Income section
        _sec('INCOME')
        _col_hdrs()
        _brow_data('Salary / Paycheck', income_salary, bg=C_GREEN)
        if income_bonus > 0:
            _brow_data('Bonus / Extra', income_bonus, bg=C_GREEN)
        _brow_data('TOTAL INCOME', total_income, is_total=True)
        total_income_brow = brow - 1
        brow += 1

        # Expenses section — categories come from categories.json
        _sec('EXPENSES')
        _col_hdrs()

        for cat in top_level:
            children = subcats.get(cat, [])
            if children:
                parent_actual = (
                    cat_actuals.get(cat, 0.0)
                    + sum(cat_actuals.get(c, 0.0) for c in children)
                )
                _brow_data(cat, parent_actual, bg=C_TBLUE)
                ws1.cell(brow - 1, 1).font = _fnt(bold=True, size=10)
                for child in children:
                    _brow_data(child, cat_actuals.get(child, 0.0), indent=True)
            else:
                _brow_data(cat, cat_actuals.get(cat, 0.0))

        # Any spend not in categories.json
        known_cats = set(all_cats)
        uncategorized = (
            cat_actuals.get('Uncategorized', 0.0)
            + sum(v for k, v in cat_actuals.items()
                  if k not in known_cats and k != 'Uncategorized')
        )
        if uncategorized > 0:
            _brow_data('Uncategorized', uncategorized)

        _brow_data('TOTAL EXPENSES', total_expenses, is_total=True)
        total_expense_brow = brow - 1
        brow += 1

        # Summary
        _sec('SUMMARY')
        net_bg = C_GREEN if net >= 0 else C_RED
        ws1.cell(brow, 1, 'Net Savings  (Income − Expenses)')
        ws1.cell(brow, 2, '')
        ws1.cell(brow, 3, round(net, 2))
        ws1.cell(brow, 4, f'=C{total_income_brow}-C{total_expense_brow}')
        for ci in range(1, 5):
            c = ws1.cell(brow, ci)
            c.fill      = _fill(net_bg)
            c.font      = _fnt(bold=True, size=11)
            c.alignment = _aln('right' if ci > 1 else 'left')
        ws1.cell(brow, 3).number_format = MONEY
        ws1.cell(brow, 4).number_format = MONEY
        ws1.row_dimensions[brow].height = 22
        brow += 1

        savings_rate = net / total_income if total_income > 0 else 0
        ws1.cell(brow, 1, 'Savings Rate')
        ws1.cell(brow, 3, savings_rate)
        ws1.cell(brow, 1).font = _fnt(italic=True, size=10)
        ws1.cell(brow, 3).number_format = PCT

        # ════════════════════════════════════════════════════════════════════
        # SHEET 2 — Transactions  (no Frequency column)
        # Date | Type | Description | Category | Amount | Notes
        # ════════════════════════════════════════════════════════════════════
        ws2 = wb.create_sheet('Transactions')

        ws2.append([f'Transactions — {label_month}'])
        ws2.merge_cells('A1:F1')
        t = ws2.cell(1, 1)
        t.fill      = _fill(C_NAVY)
        t.font      = _fnt(bold=True, color=C_WHITE, size=13)
        t.alignment = _aln('center')
        ws2.row_dimensions[1].height = 26

        tx_headers = ['Date', 'Type', 'Description', 'Category', 'Amount', 'Notes']
        tx_widths  = [13,     14,     36,             22,         14,       32     ]
        ws2.append(tx_headers)
        for i, _ in enumerate(tx_headers, 1):
            c = ws2.cell(2, i)
            c.fill      = _fill(C_BLUE)
            c.font      = _fnt(bold=True, color=C_WHITE, size=10)
            c.alignment = _aln('center')
        ws2.row_dimensions[2].height = 19
        ws2.freeze_panes = 'A3'
        for i, w in enumerate(tx_widths, 1):
            ws2.column_dimensions[get_column_letter(i)].width = w
        ws2.auto_filter.ref = f"A2:F{len(tx_rows) + 2}"

        TYPE_LABELS = {'expense': '💸 Expense', 'income': '💵 Income'}

        for ri, tx in enumerate(tx_rows, 3):
            _, tx_date, place, amount, category, label, tx_type = tx
            amount_val = round(float(amount or 0), 2)
            is_reimb   = tx_type == 'expense' and amount_val < 0
            type_label = '↩️ Reimb.' if is_reimb else TYPE_LABELS.get(tx_type, tx_type.title())

            ws2.append([str(tx_date or ''), type_label, str(place or ''),
                        str(category or ''), amount_val, ''])

            if tx_type == 'income':
                row_bg = C_GREEN
            elif is_reimb:
                row_bg = C_PURP
            else:
                row_bg = C_RED if ri % 2 == 0 else C_GREY

            for ci in range(1, 7):
                c = ws2.cell(ri, ci)
                c.fill      = _fill(row_bg)
                c.alignment = _aln('center' if ci in (1, 2) else 'left',
                                   wrap=(ci == 6))

            amt_c = ws2.cell(ri, 5)
            amt_c.number_format = MONEY
            amt_c.alignment     = _aln('right')
            if tx_type == 'income':
                amt_c.font = _fnt(color='166534', bold=True)
            elif is_reimb:
                amt_c.font = _fnt(color='5B21B6', bold=True)

        # Net row
        net_row = len(tx_rows) + 3
        ws2.append(['', '', '', 'NET', round(net, 2), ''])
        for ci in range(1, 7):
            c = ws2.cell(net_row, ci)
            c.fill = _fill(C_NAVY)
            c.font = _fnt(bold=True, color=C_WHITE, size=11)
        ws2.cell(net_row, 5).number_format = MONEY
        ws2.cell(net_row, 5).alignment = _aln('right')

        # ════════════════════════════════════════════════════════════════════
        # SHEET 3 — Category Detail
        # ════════════════════════════════════════════════════════════════════
        ws3 = wb.create_sheet('Category Detail')

        ws3.append([f'Category Detail — {label_month}'])
        ws3.merge_cells('A1:F1')
        h = ws3.cell(1, 1)
        h.fill      = _fill(C_NAVY)
        h.font      = _fnt(bold=True, color=C_WHITE, size=13)
        h.alignment = _aln('center')
        ws3.row_dimensions[1].height = 26

        cd_hdrs   = ['Category', 'Total Spent', 'Transactions', 'Avg / Transaction',
                     '% of Expenses', '% of Income']
        cd_widths = [28, 16, 14, 20, 16, 14]
        ws3.append(cd_hdrs)
        for i, _ in enumerate(cd_hdrs, 1):
            c = ws3.cell(2, i)
            c.fill      = _fill(C_BLUE)
            c.font      = _fnt(bold=True, color=C_WHITE, size=10)
            c.alignment = _aln('center')
        ws3.row_dimensions[2].height = 19
        ws3.freeze_panes = 'A3'
        ws3.auto_filter.ref = 'A2:F2'
        for i, w in enumerate(cd_widths, 1):
            ws3.column_dimensions[get_column_letter(i)].width = w

        cat_agg: dict = {}
        for tx in tx_rows:
            _, _, _, amount, category, _, tx_type = tx
            amount_val = round(float(amount or 0), 2)
            if tx_type != 'expense' or amount_val < 0:
                continue
            cat = str(category or 'Uncategorized')
            if cat not in cat_agg:
                cat_agg[cat] = {'total': 0.0, 'count': 0}
            cat_agg[cat]['total'] += amount_val
            cat_agg[cat]['count'] += 1

        sorted_cats = sorted(cat_agg.items(), key=lambda x: -x[1]['total'])

        for i, (cat, data) in enumerate(sorted_cats):
            rn      = i + 3
            avg     = data['total'] / data['count'] if data['count'] else 0
            pct_exp = data['total'] / total_expenses if total_expenses > 0 else 0
            pct_inc = data['total'] / total_income   if total_income   > 0 else ''
            ws3.append([cat, round(data['total'], 2), data['count'], round(avg, 2),
                        pct_exp, pct_inc])
            bg = C_GREY if i % 2 == 0 else C_WHITE
            for ci in range(1, 7):
                ws3.cell(rn, ci).fill = _fill(bg)
            for ci in (2, 4):
                ws3.cell(rn, ci).number_format = MONEY
                ws3.cell(rn, ci).alignment = _aln('right')
            ws3.cell(rn, 3).alignment = _aln('right')
            for ci in (5, 6):
                if pct_inc != '' or ci == 5:
                    ws3.cell(rn, ci).number_format = PCT
                    ws3.cell(rn, ci).alignment = _aln('right')

        # Grand-total row
        gt_row   = len(sorted_cats) + 3
        gt_total = sum(d['total'] for _, d in sorted_cats)
        gt_count = sum(d['count'] for _, d in sorted_cats)
        ws3.append(['TOTAL EXPENSES', round(gt_total, 2), gt_count, '',
                    1.0 if sorted_cats else '', ''])
        for ci in range(1, 7):
            c = ws3.cell(gt_row, ci)
            c.fill = _fill(C_NAVY)
            c.font = _fnt(bold=True, color=C_WHITE, size=11)
        ws3.cell(gt_row, 2).number_format = MONEY
        ws3.cell(gt_row, 2).alignment = _aln('right')
        ws3.cell(gt_row, 5).number_format = PCT

        # ── Serialise ─────────────────────────────────────────────────────────
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"budget_{month or 'all'}.xlsx"
        return StreamingResponse(
            buf,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename="{filename}"'},
        )

    except Exception as exc:
        logging.exception('Excel export failed')
        return JSONResponse(status_code=500, content={'error': str(exc)})


        # Title banner
        ws_tx.append([f'Personal Budget — {label_month}'])
        ws_tx.merge_cells('A1:G1')
        title_cell = ws_tx.cell(1, 1)
        title_cell.fill = HDR_FILL
        title_cell.font = Font(bold=True, color='FFFFFF', size=14)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws_tx.row_dimensions[1].height = 28

        # Column headers (row 2)
        tx_headers = ['Date', 'Type', 'Description', 'Category', 'Amount', 'Frequency', 'Notes']
        tx_widths  = [13,     14,     36,             22,         14,       14,           30    ]
        ws_tx.append(tx_headers)
        for i, _ in enumerate(tx_headers, 1):
            c = ws_tx.cell(2, i)
            c.fill = SUB_FILL; c.font = SUB_FONT
            c.alignment = Alignment(horizontal='center', vertical='center')
        ws_tx.row_dimensions[2].height = 20
        ws_tx.freeze_panes = 'A3'
        for i, w in enumerate(tx_widths, 1):
            ws_tx.column_dimensions[get_column_letter(i)].width = w
        ws_tx.auto_filter.ref = f"A2:G{len(tx_rows) + 2}"

        FREQ_MAP = {
            'recurring':     'Regular',
            'one-time':      'One-Time',
            'bonus':         'Bonus',
            'reimbursement': 'Reimbursement',
        }
        TYPE_LABELS = {'expense': '💸 Expense', 'income': '💵 Income'}

        for row_num, r in enumerate(tx_rows, 3):
            _, tx_date, place, amount, category, label, tx_type, report_month = r
            amount_val = round(float(amount or 0), 2)
            is_reimb   = tx_type == 'expense' and amount_val < 0
            type_label = '↩️ Reimb.' if is_reimb else TYPE_LABELS.get(tx_type, tx_type.title())
            freq       = FREQ_MAP.get(str(label or '').lower(), str(label or '').title() or 'Regular')

            ws_tx.append([
                str(tx_date or ''),
                type_label,
                str(place or ''),
                str(category or ''),
                amount_val,
                freq,
                '',   # Notes — blank for user to fill
            ])

            if tx_type == 'income':
                row_fill = INCOME_FILL
            elif is_reimb:
                row_fill = REIMB_FILL
            else:
                row_fill = EXPENSE_FILL if row_num % 2 == 0 else ALT_FILL

            for col in range(1, 8):
                c = ws_tx.cell(row_num, col)
                c.fill = row_fill
                c.alignment = Alignment(vertical='center', wrap_text=(col == 7))

            amt_cell = ws_tx.cell(row_num, 5)
            amt_cell.number_format = MONEY_FMT
            amt_cell.alignment = Alignment(horizontal='right', vertical='center')
            if tx_type == 'income':
                amt_cell.font = Font(color='166534', bold=True)
            elif is_reimb:
                amt_cell.font = Font(color='5B21B6', bold=True)

        # Totals row at bottom of ledger
        total_row = len(tx_rows) + 3
        ws_tx.append(['', '', '', 'TOTAL', round(total_income - total_expenses + total_reimb, 2), '', ''])
        for col in range(1, 8):
            c = ws_tx.cell(total_row, col)
            c.fill = TOTAL_FILL; c.font = TOTAL_FONT
        ws_tx.cell(total_row, 5).number_format = MONEY_FMT
        ws_tx.cell(total_row, 5).alignment = Alignment(horizontal='right', vertical='center')

        # ── Sheet 2: Category Breakdown ───────────────────────────────────────
        ws_cat = wb.create_sheet('Category Breakdown')
        ws_cat.append([f'Category Breakdown — {label_month}'])
        ws_cat.merge_cells('A1:F1')
        c1 = ws_cat.cell(1, 1)
        c1.fill = HDR_FILL; c1.font = Font(bold=True, color='FFFFFF', size=13)
        c1.alignment = Alignment(horizontal='center', vertical='center')
        ws_cat.row_dimensions[1].height = 26

        cat_headers = ['Category', 'Total Spent', 'Transactions', 'Avg / Transaction', '% of Expenses', '% of Income']
        cat_widths  = [28,         16,             14,             20,                  16,               14           ]
        _hdr(ws_cat, cat_headers, cat_widths)
        ws_cat.freeze_panes = 'A3'

        cat_agg: dict = {}
        for r in tx_rows:
            _, _, _, amount, category, _, tx_type, _ = r
            amount_val = round(float(amount or 0), 2)
            if tx_type != 'expense' or amount_val < 0:
                continue
            cat = str(category or 'Uncategorized')
            if cat not in cat_agg:
                cat_agg[cat] = {'total': 0.0, 'count': 0}
            cat_agg[cat]['total'] += amount_val
            cat_agg[cat]['count'] += 1

        sorted_cats = sorted(cat_agg.items(), key=lambda x: -x[1]['total'])
        pct_income_avail = total_income > 0

        for i, (cat, data) in enumerate(sorted_cats):
            row_num = i + 3   # row 1=title, row 2=headers
            avg = data['total'] / data['count'] if data['count'] else 0
            pct_exp = data['total'] / total_expenses if total_expenses > 0 else 0
            pct_inc = data['total'] / total_income   if pct_income_avail else ''
            ws_cat.append([cat, round(data['total'], 2), data['count'], round(avg, 2), pct_exp, pct_inc])
            fill = ALT_FILL if i % 2 == 0 else PatternFill('solid', fgColor='FFFFFF')
            for col in range(1, 7):
                ws_cat.cell(row_num, col).fill = fill
            for col in (2, 4):
                ws_cat.cell(row_num, col).number_format = MONEY_FMT
                ws_cat.cell(row_num, col).alignment = Alignment(horizontal='right')
            for col in (5, 6):
                if pct_income_avail or col == 5:
                    ws_cat.cell(row_num, col).number_format = PCT_FMT
                    ws_cat.cell(row_num, col).alignment = Alignment(horizontal='right')

        # Grand-total row
        grand_total = sum(d['total'] for _, d in sorted_cats)
        grand_count = sum(d['count'] for _, d in sorted_cats)
        gt_row = len(sorted_cats) + 3
        ws_cat.append(['TOTAL EXPENSES', round(grand_total, 2), grand_count, '', 1.0 if sorted_cats else '', ''])
        for col in range(1, 7):
            c = ws_cat.cell(gt_row, col)
            c.fill = TOTAL_FILL; c.font = TOTAL_FONT
        ws_cat.cell(gt_row, 2).number_format = MONEY_FMT
        ws_cat.cell(gt_row, 2).alignment = Alignment(horizontal='right')
        ws_cat.cell(gt_row, 5).number_format = PCT_FMT

        # ── Sheet 3: Budget Overview ──────────────────────────────────────────
        ws_ov = wb.create_sheet('Budget Overview')

        def _ov_title(ws, text, row, cols='A:C'):
            ws.cell(row, 1, text).fill = HDR_FILL
            ws.cell(row, 1).font       = Font(bold=True, color='FFFFFF', size=12)
            ws.cell(row, 1).alignment  = Alignment(horizontal='left', vertical='center')
            ws.merge_cells(f'A{row}:{cols.split(":")[1]}{row}')
            ws.row_dimensions[row].height = 22

        def _ov_sub(ws, text, row):
            ws.cell(row, 1, text).fill = SUB_FILL
            ws.cell(row, 1).font       = SUB_FONT
            ws.cell(row, 1).alignment  = Alignment(horizontal='left', vertical='center')
            ws.merge_cells(f'A{row}:E{row}')
            ws.row_dimensions[row].height = 18

        def _ov_row(ws, row, lbl, actual, target, fill=None):
            variance = actual - target if target else 0
            status   = '✅ On Track' if variance <= 0 else '⚠️ Over Budget'
            row_fill = (fill or (OK_FILL if variance <= 0 else OVER_FILL))
            ws.cell(row, 1, lbl).alignment = Alignment(vertical='center')
            ws.cell(row, 2, round(actual,   2)).number_format = MONEY_FMT
            ws.cell(row, 3, round(target,   2)).number_format = MONEY_FMT
            ws.cell(row, 4, round(variance, 2)).number_format = MONEY_FMT
            ws.cell(row, 5, status)
            for col in range(1, 6):
                c = ws_ov.cell(row, col)
                c.fill = row_fill
                c.alignment = Alignment(vertical='center',
                                        horizontal='right' if col in (2, 3, 4) else 'left')
            ws.row_dimensions[row].height = 18
            return variance

        for col, w in zip('ABCDE', [30, 16, 16, 16, 16]):
            ws_ov.column_dimensions[col].width = w

        # Section 1: Income & Summary
        _ov_title(ws_ov, f'Budget Overview — {label_month}', 1, 'A:E')
        r = 2
        for col, hdr in enumerate(['', 'Actual', 'Target / Guide', 'Variance', 'Status'], 1):
            c = ws_ov.cell(r, col, hdr)
            c.fill = SUB_FILL; c.font = SUB_FONT
            c.alignment = Alignment(horizontal='center' if col > 1 else 'left', vertical='center')
        ws_ov.row_dimensions[r].height = 18
        r += 1

        # Income row
        ws_ov.cell(r, 1, '💵 Total Income').font = Font(bold=True)
        ws_ov.cell(r, 2, round(total_income, 2)).number_format = MONEY_FMT
        ws_ov.cell(r, 3, '—'); ws_ov.cell(r, 4, '—'); ws_ov.cell(r, 5, '—')
        for col in range(1, 6):
            ws_ov.cell(r, col).fill = INCOME_FILL
            ws_ov.cell(r, col).alignment = Alignment(
                vertical='center', horizontal='right' if col == 2 else 'left')
        ws_ov.row_dimensions[r].height = 18
        r += 1

        # Net savings row
        savings_target = round(total_income * 0.20, 2)
        ws_ov.cell(r, 1, '💰 Net Savings').font = Font(bold=True)
        ws_ov.cell(r, 2, round(net_savings, 2)).number_format = MONEY_FMT
        ws_ov.cell(r, 3, savings_target).number_format = MONEY_FMT
        savings_rate = net_savings / total_income if total_income > 0 else 0
        ws_ov.cell(r, 4, f'{savings_rate:.1%} of income')
        ws_ov.cell(r, 5, '✅ On Track' if savings_rate >= 0.20 else '⚠️ Below 20% target')
        sf = OK_FILL if savings_rate >= 0.20 else WARN_FILL
        for col in range(1, 6):
            ws_ov.cell(r, col).fill = sf
            ws_ov.cell(r, col).alignment = Alignment(
                vertical='center', horizontal='right' if col in (2, 3) else 'left')
        ws_ov.row_dimensions[r].height = 18
        r += 2   # blank spacer

        # Section 2: 50/30/20 Rule
        _ov_sub(ws_ov, '50 / 30 / 20 Rule Analysis  (based on total income)', r)
        r += 1

        NEEDS_CATS   = {'Housing', 'Rent', 'Mortgage', 'Utilities', 'Electric', 'Natural Gas',
                        'Water/Sewer', 'Internet/Cable', 'Groceries', 'Insurance', 'Health',
                        'Medical', 'Transportation', 'Gas/Fuel', 'Auto Maintenance', 'Childcare'}
        WANTS_CATS   = {'Dining', 'Entertainment', 'Shopping', 'Subscriptions', 'Travel',
                        'Vacation', 'Hobbies', 'Gym', 'Personal Care', 'Gifts & Charity',
                        'Gifts & Donations', 'Clothing', 'Electronics'}
        SAVINGS_CATS = {'Investment', 'Investment Transfer', 'Savings', 'Retirement', 'Emergency Fund'}

        buckets: dict = {'Needs': 0.0, 'Wants': 0.0, 'Savings & Investments': 0.0, 'Other': 0.0}
        for r_tx in tx_rows:
            _, _, _, amount, category, _, tx_type, _ = r_tx
            amount_val = round(float(amount or 0), 2)
            if tx_type != 'expense' or amount_val < 0:
                continue
            cat = str(category or '').strip()
            if cat in NEEDS_CATS:
                buckets['Needs'] += amount_val
            elif cat in WANTS_CATS:
                buckets['Wants'] += amount_val
            elif cat in SAVINGS_CATS:
                buckets['Savings & Investments'] += amount_val
            else:
                buckets['Other'] += amount_val

        bucket_targets = {
            'Needs':                 round(total_income * 0.50, 2),
            'Wants':                 round(total_income * 0.30, 2),
            'Savings & Investments': round(total_income * 0.20, 2),
            'Other':                 0.0,
        }
        BUCKET_EMOJI = {
            'Needs': '🏠', 'Wants': '🎉',
            'Savings & Investments': '📈', 'Other': '📦',
        }

        for col, hdr in enumerate(['Category', 'Actual', 'Target (50/30/20)', 'Variance', 'Status'], 1):
            c = ws_ov.cell(r, col, hdr)
            c.fill = SUB_FILL; c.font = SUB_FONT
            c.alignment = Alignment(horizontal='center' if col > 1 else 'left', vertical='center')
        ws_ov.row_dimensions[r].height = 18
        r += 1

        for bucket, actual in buckets.items():
            target  = bucket_targets[bucket]
            label_b = f"{BUCKET_EMOJI[bucket]} {bucket}"
            _ov_row(ws_ov, r, label_b, actual, target)
            r += 1

        r += 1   # blank spacer

        # Section 3: Category detail
        _ov_sub(ws_ov, 'Expense Breakdown by Category', r)
        r += 1

        for col, hdr in enumerate(['Category', 'Total Spent', '% of Expenses', '% of Income', ''], 1):
            c = ws_ov.cell(r, col, hdr)
            c.fill = SUB_FILL; c.font = SUB_FONT
            c.alignment = Alignment(horizontal='center' if col > 1 else 'left', vertical='center')
        ws_ov.row_dimensions[r].height = 18
        r += 1

        for i, (cat, data) in enumerate(sorted_cats):
            pct_exp = data['total'] / total_expenses if total_expenses > 0 else 0
            pct_inc = data['total'] / total_income   if total_income   > 0 else 0
            fill = ALT_FILL if i % 2 == 0 else PatternFill('solid', fgColor='FFFFFF')
            ws_ov.cell(r, 1, cat).fill = fill
            ws_ov.cell(r, 2, round(data['total'], 2)).number_format = MONEY_FMT
            ws_ov.cell(r, 3, pct_exp).number_format = PCT_FMT
            ws_ov.cell(r, 4, pct_inc).number_format = PCT_FMT
            for col in range(1, 5):
                ws_ov.cell(r, col).fill = fill
                ws_ov.cell(r, col).alignment = Alignment(
                    vertical='center', horizontal='right' if col > 1 else 'left')
            r += 1

        # ── Serialise ─────────────────────────────────────────────────────────
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"budget_{month or 'all'}.xlsx"
        return StreamingResponse(
            buf,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename="{filename}"'},
        )

    except Exception as exc:
        logging.exception("Excel export failed")
        return JSONResponse(status_code=500, content={'error': str(exc)})
